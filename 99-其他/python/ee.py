from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.expected_conditions import presence_of_element_located, visibility_of_element_located
import time
from bs4 import BeautifulSoup
import openpyxl

# 说明： 使用Selenium驱动Chrome浏览器，并在脚本结束后自动退出
# 注意： 需要安装Selenium驱动包，如：pip install selenium
# 注意： 需要安装openpyxl包，如：pip install openpyxl
# 注意： 需要安装BeautifulSoup4包，如：pip install beautifulsoup4
# 注意： 需要安装requests包，如：pip install requests
# 注意： 需要安装对应版本的chromedriver包，如：https://chromedriver.storage.googleapis.com/index.html 。 需要配置环境变量，加到PATH中

# 手动登录，登录后按Enter键继续脚本执行
# 读取Excel a列的每个数据，拼接到url中，然后访问该url，获取页面源代码，然后解析页面，获取B列的值，然后更新Excel文件

# 初始化选项，使浏览器在脚本结束后保持打开状态
options = Options()
options.add_experimental_option("detach", True)

# 初始化Chrome浏览器实例
driver = webdriver.Chrome(options=options)

# 访问登录页面
driver.get("https://vsp.jd.com/passport/login")

# 这里假设需要用户手动登录，可替换为自动提交表单逻辑
print("请手动登录后按Enter键继续...")
input("按Enter键继续脚本执行...")

# Excel文件路径
excel_path = 'aaa.xlsx'
# SKU查询URL模板
url_template = 'https://vsp.jd.com/new/sku/{}'

# 打开Excel文件
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# 遍历Excel中的每一行，从第二行开始
for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    sku = str(row[0])  # 假设A列是SKU
    current_b_value = row[1]  # 假设B列是当前状态

    # 构造完整URL并访问
    full_url = url_template.format(sku)
    driver.get(full_url)

    # 确保页面完全加载
    WebDriverWait(driver, 10).until(visibility_of_element_located((By.TAG_NAME, "body")))
    time.sleep(2) 

    # 获取页面源码并解析
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, "html.parser")

    # 检查页面是否包含特定字符串“加入采购单”，并更新Excel
    if "加入采购单" in soup.text:
        print(f"SKU {sku} 页面包含'加入采购单'，更新Excel为'已在'。")
        sheet.cell(row=index, column=2, value='已在')
    elif "您不能查看该商品，因为该商品不在您的商品池中" in soup.text:
        print(f"SKU {sku} 页面包含'您不能查看该商品，因为该商品不在您的商品池中'，更新Excel为'不在商品池中'。")
        sheet.cell(row=index, column=2, value='不在商品池中')
    else:
        print(f"SKU {sku} 页面不含'加入采购单'，更新Excel为'不包含'。")
        sheet.cell(row=index, column=2, value='不包含')

    # 保存每次更新
    workbook.save(excel_path)

# 这里可以继续其他操作...
# 注意：脚本执行完毕后，浏览器会保持打开状态，根据需要手动关闭